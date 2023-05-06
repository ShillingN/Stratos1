from .entities import AutoChange, PlanOrder
from db import Session
from threading import Timer
from errors import *
import datetime
import random
import orders
import tools
import utils
import pairs
import users
import traceback
import files
import openpyxl


def get_plans_orders(date):
    session = Session()
    plan_orders = session.query(PlanOrder).all()

    result = []

    for plan_order in plan_orders:
        if utils.compare_days(plan_order.date, date):
            result.append(plan_order)
    return result



def check_plans():
    session = Session()
    try:
        setting = AutoChange()
        now = datetime.datetime.now()
        today = now.replace(hour=0, minute=0, second=0, microsecond=0)
        plan_orders = session.query(PlanOrder).filter(PlanOrder.date > today).all()

        if len(plan_orders) == 0:
            generate_orders()
        else:
            if setting.status is True:
                for plan_order in plan_orders:
                    if plan_order.status is False and utils.compare_date(now, plan_order.date):
                        user = get_user()
                        value = random.randint(setting.min_value, setting.max_value)
                        pair, value = get_pair(value)
                        order = orders.api.create_order(user.email, pair, value, 'Автообмен', ref_id=setting.ref_id, is_auto=True)
                        orders.api.change_order_status(order.id, 3)
                        plan_order.status = True
    except:
        print(traceback.format_exc())
    finally:
        session.commit()
        session.close()
        Timer(60, check_plans).start()


def generate_orders():
    session = Session()
    setting = AutoChange()

    now = datetime.datetime.now()
    time = datetime.datetime(now.year, now.month, now.day, setting.start_hour)
    end_time = datetime.datetime(now.year, now.month, now.day, setting.end_hour)

    while time < end_time:
        value = random.randint(setting.min_value, setting.max_value)
        plan_order = PlanOrder(date=time, value=value, ref_id=setting.ref_id, status=False)
        session.add(plan_order)

        time += datetime.timedelta(minutes=random.randint(setting.min_minutes, setting.max_minutes))
    session.commit()


def get_user():
    setting = AutoChange()
    emails = get_users_list()

    email_index = setting.user_id + 1
    if email_index >= len(emails):
        email_index = 0

    setting.user_id = email_index
    setting.save()
    return emails[email_index]


def get_users_list():
    wb = openpyxl.load_workbook('{}/emails.xlsx'.format(utils.get_script_dir()))
    sheet = wb.active
    emails = []
    for i in sheet.iter_rows():
        emails.append(i[0].value)
    return emails


def get_pair(value):
    btc = tools.api.get_tool_by_name('BTC', None)
    eth = tools.api.get_tool_by_name('ETH', None)
    usdt = tools.api.get_tool_by_name('USDT', 'TRC20')
    sber = tools.api.get_tool_by_name('Сбербанк', None)
    pairs_list = [
        pairs.api.get_pair_by_tools(btc.id, sber.id),
        pairs.api.get_pair_by_tools(eth.id, sber.id),
        pairs.api.get_pair_by_tools(usdt.id, sber.id),
    ]

    pair = random.choice(pairs_list)
    pair.get_cost()
    value = 1 / pair.get_cost() * value
    return random.choice(pairs_list), value


def upload_file(file_uuid):
    file = files.api.get_file(file_uuid)
    if file is None:
        raise IncorrectDataValue('Файл не найден')

    wb = openpyxl.load_workbook('{}/storage/{}.{}'.format(utils.get_script_dir(), file.uuid, file.extension))
    sheet = wb.active
    emails = []
    for i in sheet.iter_rows():
        emails.append(i[0].value)

    wb = openpyxl.workbook.Workbook()
    sheet = wb.create_sheet('Лист 1', 0)
    index = 1
    for email in emails:
        sheet.cell(row=index, column=1).value = email
        index += 1
    wb.save('{}/emails.xlsx'.format(utils.get_script_dir()))